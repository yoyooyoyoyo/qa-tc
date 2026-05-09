from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Iterable

from app.schemas.testcase import TestCase


TESTCASE_EXPORT_HEADERS = [
    "testcase_id",
    "requirement_id",
    "perspective",
    "priority",
    "test_type",
    "category",
    "severity",
    "automation_candidate",
    "title",
    "test_screen",
    "preconditions",
    "steps",
    "test_data",
    "expected_result",
    "related_risks",
    "traceability",
    "source_quote",
    "source_policy",
    "related_policy",
    "generation_scope",
    "risk_basis",
    "omit_reason",
    "quality_warnings",
    "notes",
]

TESTCASE_XLSX_HEADERS = [
    "no.",
    "title",
    "테스트 진행 화면",
    "precondition",
    "steps",
    "expected result",
    "test-result",
    "comment",
    "issue ticket",
]

TEST_RESULT_OPTIONS = ["pass", "fail", "blocked", "fixed"]


def testcases_to_tsv(testcases: Iterable[TestCase]) -> str:
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=TESTCASE_EXPORT_HEADERS,
        delimiter="\t",
        lineterminator="\n",
    )
    writer.writeheader()

    for testcase in testcases:
        writer.writerow(_flatten_testcase(testcase))

    return output.getvalue()


def testcases_to_xlsx(testcases: Iterable[TestCase]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx export.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "testcases"
    worksheet.append(TESTCASE_XLSX_HEADERS)

    for idx, testcase in enumerate(testcases, start=1):
        row = _flatten_testcase_for_xlsx(idx, testcase)
        worksheet.append([row[header] for header in TESTCASE_XLSX_HEADERS])

    _style_testcase_worksheet(
        worksheet,
        header_font=Font(bold=True, color="FFFFFF"),
        header_fill=PatternFill("solid", fgColor="4F81BD"),
        wrap_alignment=Alignment(wrap_text=True, vertical="top"),
    )
    _add_test_result_dropdown(worksheet, DataValidation)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _flatten_testcase(testcase: TestCase) -> dict[str, str]:
    return {
        "testcase_id": testcase.testcase_id,
        "requirement_id": testcase.requirement_id,
        "perspective": testcase.perspective,
        "priority": testcase.priority,
        "test_type": testcase.test_type,
        "category": testcase.category,
        "severity": testcase.severity,
        "automation_candidate": str(testcase.automation_candidate),
        "title": testcase.title,
        "test_screen": testcase.test_screen,
        "preconditions": "\n".join(testcase.preconditions),
        "steps": "\n".join(testcase.steps),
        "test_data": testcase.test_data,
        "expected_result": testcase.expected_result,
        "related_risks": "\n".join(testcase.related_risks),
        "traceability": "\n".join(testcase.traceability),
        "source_quote": testcase.source_quote,
        "source_policy": "\n".join(testcase.source_policy),
        "related_policy": "\n".join(testcase.related_policy),
        "generation_scope": testcase.generation_scope,
        "risk_basis": "\n".join(testcase.risk_basis),
        "omit_reason": testcase.omit_reason,
        "quality_warnings": "\n".join(testcase.quality_warnings),
        "notes": testcase.notes,
    }


def _flatten_testcase_for_xlsx(idx: int, testcase: TestCase) -> dict[str, str]:
    return {
        "no.": str(idx),
        "title": testcase.title,
        "테스트 진행 화면": testcase.test_screen,
        "precondition": "\n".join(testcase.preconditions),
        "steps": "\n".join(testcase.steps),
        "expected result": testcase.expected_result,
        "test-result": "",
        "comment": "",
        "issue ticket": "",
    }


def _style_testcase_worksheet(
    worksheet,
    *,
    header_font,
    header_fill,
    wrap_alignment,
) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 8,
        "B": 36,
        "C": 28,
        "D": 30,
        "E": 48,
        "F": 42,
        "G": 18,
        "H": 34,
        "I": 24,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment


def _add_test_result_dropdown(worksheet, data_validation_cls) -> None:
    options = ",".join(TEST_RESULT_OPTIONS)
    validation = data_validation_cls(
        type="list",
        formula1=f'"{options}"',
        allow_blank=True,
    )
    validation.error = "pass, fail, blocked, fixed 중 하나를 선택하세요."
    validation.errorTitle = "Invalid test result"
    validation.prompt = "pass, fail, blocked, fixed 중 선택"
    validation.promptTitle = "test-result"

    worksheet.add_data_validation(validation)
    max_row = max(worksheet.max_row, 1000)
    validation.add(f"G2:G{max_row}")
